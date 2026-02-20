"""
Excel parser for IO sheets and commissioning spreadsheets.

Uses openpyxl in read-only mode for memory efficiency.
Normalizes 20+ column aliases to canonical names.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from app.models.project_models import IORecord

logger = logging.getLogger(__name__)

# Column alias normalisation map (lowercase → canonical)
_COL_MAP: dict[str, str] = {
    # slot / point
    "slot": "slot", "point": "slot", "io point": "slot", "io_point": "slot",
    "channel": "slot", "address": "slot",
    # rack
    "rack": "rack", "chassis": "rack", "frame": "rack",
    # module / card
    "module": "module", "card": "module", "card type": "module",
    "device type": "module", "module type": "module", "i/o module": "module",
    # description
    "description": "description", "tag description": "description",
    "signal description": "description", "label": "description",
    "comment": "description", "function": "description",
    # tag name
    "tag": "tag_name", "tag name": "tag_name", "plc tag": "tag_name",
    "address tag": "tag_name", "symbol": "tag_name",
}

# Rows whose first cell looks like a section header are skipped
_SKIP_KEYWORDS = {"no.", "no", "#", "item", "ref", "s/n", "sl no"}


@dataclass
class ExcelParseResult:
    io_rows: list[IORecord] = field(default_factory=list)
    source_file: str = ""
    sheets_parsed: int = 0
    warnings: list[str] = field(default_factory=list)


def parse(file_path: str | Path) -> ExcelParseResult:
    """
    Parse all sheets in an Excel file, return IO rows.
    Never raises — returns warnings in result.
    """
    try:
        import openpyxl
    except ImportError:
        return ExcelParseResult(
            source_file=str(file_path),
            warnings=["openpyxl not installed — Excel parsing skipped."],
        )

    path = Path(file_path)
    result = ExcelParseResult(source_file=str(path))

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        result.warnings.append(f"Cannot open {path.name}: {exc}")
        return result

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Find header row — first row with recognisable column names
        header_idx, col_map = _find_header(rows)
        if not col_map:
            result.warnings.append(
                f"{path.name} sheet='{sheet_name}': No recognisable IO columns found."
            )
            continue

        result.sheets_parsed += 1
        for row in rows[header_idx + 1:]:
            io = _parse_row(row, col_map, str(path))
            if io:
                result.io_rows.append(io)

    wb.close()
    logger.debug("Excel parsed: %s → %d IO rows", path.name, len(result.io_rows))
    return result


def _find_header(rows: list[tuple]) -> tuple[int, dict[int, str]]:
    """Scan first 10 rows for the header row. Returns (row_index, {col_idx→canonical})."""
    for i, row in enumerate(rows[:10]):
        col_map: dict[int, str] = {}
        for j, cell in enumerate(row):
            if cell is None:
                continue
            key = str(cell).strip().lower().rstrip(":").rstrip("*")
            canonical = _COL_MAP.get(key)
            if canonical:
                col_map[j] = canonical
        if len(col_map) >= 2:
            return i, col_map
    return 0, {}


def _parse_row(row: tuple, col_map: dict[int, str], source: str) -> IORecord | None:
    data: dict[str, Any] = {}
    for idx, canonical in col_map.items():
        if idx < len(row) and row[idx] is not None:
            data[canonical] = str(row[idx]).strip()

    # Skip empty rows and section headers
    slot = data.get("slot", "").strip()
    if not slot or slot.lower() in _SKIP_KEYWORDS:
        return None

    return IORecord(
        slot=slot,
        rack=data.get("rack", ""),
        module=data.get("module", ""),
        description=data.get("description", ""),
        tag_name=data.get("tag_name", ""),
        source_file=source,
    )
